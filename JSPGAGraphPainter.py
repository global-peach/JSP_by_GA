import numpy as np
import matplotlib.pyplot as plt
from Decode_for_JSP import Decode
import JSPGA
from model import HFSPGAInput
import openpyxl

class JSPGAGraphPainter:
    def main(self,inputParam: HFSPGAInput):
        ga = JSPGA.GA()
        result = ga.start(inputParam)
        self.Gantt(result.Decode, inputParam)

        # for i in range(len(result.CHS)):
        #     d = Decode(result.Decode.J, inputParam, result.Decode.M_num)
        #     d.Decode_1(result.CHS[i], result.LenChromo)
        #     self.Gantt(d, inputParam)

        x = np.linspace(0, len(result.BestFit), len(result.BestFit))
        plt.rcParams['font.sans-serif']=['Microsoft YaHei']
        plt.plot(x, result.BestFit, x, result.WorstFit, x, result.AvgFit,'-k')
        plt.title('Best, Worst, Average Scheduling time Each iterator')
        plt.ylabel('Scheduling time')
        plt.xlabel('Number of iterations')
        plt.show()

    def Gantt(self, Decode: Decode, inputParam: HFSPGAInput):
        Machines = Decode.Machines
        plt.figure(figsize=(16, 9))
        plt.rcParams['font.sans-serif']=['Microsoft YaHei']
        M = ['red', 'blue', 'yellow', 'orange', 'green', 'palegoldenrod', 'purple', 'pink', 'Thistle', 'Magenta',
             'SlateBlue', 'RoyalBlue', 'Cyan', 'Aqua', 'floralwhite', 'ghostwhite', 'goldenrod', 'mediumslateblue',
             'navajowhite',
             'navy', 'sandybrown', 'moccasin']
        totalEnd = 0
        process_data: list[list] = [
            ['机器', '工件', '工序', '开始加工时间', '结束加工时间', '耗时', '砝码', '机器等待时长', '工件等待时长']
        ]
        for i in range(len(Machines)):
            Machine=Machines[i]
            Start_time=Machine.O_start
            End_time=Machine.O_end
            if Machine.End_time > totalEnd:
                totalEnd = Machine.End_time
            for i_1 in range(len(End_time)):
                jobIndex = Machine.assigned_task[i_1][0] - 1
                processIndex = Machine.assigned_task[i_1][1] - 1
                weightName = ''
                if i == 6 or i == 7:
                    weightIndex = Machine.worker_for_task[i_1]
                    weightName = inputParam.ProcessingWeight[jobIndex][weightIndex]
                    jobName = '{0}-{1}'.format(jobIndex + 1, weightName)
                else:
                    jobName = str(jobIndex + 1)
                # plt.barh(i,width=End_time[i_1]-Start_time[i_1],height=0.8,left=Start_time[i_1],\
                #          color=M[Machine.assigned_task[i_1][0]],edgecolor='black')
                # plt.text(x=Start_time[i_1]+0.1,y=i,s=Machine.assigned_task[i_1])
                start = Start_time[i_1]
                end = End_time[i_1]
                machineWaitTime = 0
                jobWaitTime = 0
                if i_1 > 0:
                    machineWaitTime = start - End_time[i_1 - 1]
                if processIndex > 0:
                    jobWaitTime = start - Decode.Jobs[jobIndex].J_end[processIndex - 1]
                process_data.append([
                    inputParam.WorkStationNameList[i],#机器
                    inputParam.JobNameList[jobIndex],#工件
                    inputParam.ProcessNameList[processIndex],#工序
                    start,#开始加工时间
                    end,#结束加工时间
                    end - start,#耗时
                    weightName,#砝码
                    machineWaitTime,#机器等待时长
                    jobWaitTime,#工件等待时长
                ])
                plt.barh(i, width=end - start, height=0.8, left=start, \
                         color=M[len(M) - 1 - jobIndex], edgecolor='black')
                plt.text(x=start + (end - start)/2-0.5, y=i, s=jobName)
        workbook = openpyxl.Workbook()
        if len(workbook.sheetnames) == 0:
            worksheet = workbook.create_sheet(title='总耗时' + str(end))
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        for row in range(len(process_data)):
            for column in range(len(process_data[row])):
                cell = worksheet.cell(row=row + 1, column = column + 1)
                cell.value = process_data[row][column]
        workbook.save(str(totalEnd) + '加工过程.xlsx')
        plt.yticks(np.arange(i + 1), inputParam.WorkStationNameList)
        plt.title('Scheduling Gantt chart')
        plt.ylabel('Work Center')
        plt.xlabel('Time(min)')
        plt.savefig(str(totalEnd), dpi=300)
        plt.close('all')
        #plt.show()

if __name__=='__main__':
    from Total1214 import input
    painter = JSPGAGraphPainter()
    inputParam = JSPGA.HFSPGAInput(input)
    painter.main(inputParam)